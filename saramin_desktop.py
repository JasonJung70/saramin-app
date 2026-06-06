# -*- coding: utf-8 -*-
"""
사람인 채용정보 검색기 v3.0
대상: Windows 10/11 64비트
제작: 정병남 취업코디 (50+재단 동부캠퍼스)
유효기간: 2026.12.30
"""
import sys, os, threading, webbrowser, json, datetime
import urllib.request, urllib.parse
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# ══════════════════════════════════════
# 설정
# ══════════════════════════════════════
APP_TITLE   = "채용정보 검색기 — 50+재단 동부캠퍼스"
PASSWORD    = "50pl-jason"
EXPIRE_DATE = datetime.date(2026, 12, 30)
API_KEY     = "Bi0yZ1iLm5pleggEpjdIVOfsgmyozIomwihPf0gAl7oppNl7ZWIK"
PROXY_URL   = "https://saramin-proxy.jasonjung70.workers.dev/job-search"
BATCH_SIZE  = 100   # API 1회 최대
PAGE_SIZE   = 30    # 화면 표시 단위

NOTICE_MSG = (
    "이 프로그램은 50+재단 동부캠퍼스 정병남 취업코디가\n"
    "사람인의 협조를 제공받아 제작하였으며,\n"
    "프로그램은 26년 12월말에 종료됩니다.\n\n"
    "1회에 최대 100건, 1일 10회 사용으로 제한됩니다."
)

# ── 지역 코드 (전국) ──
LOC_MAP = {
    "전국": "",
    # 서울
    "서울 전체":"101000","강남구":"101010","강동구":"101020",
    "강북구":"101030","강서구":"101040","관악구":"101050",
    "광진구":"101060","구로구":"101070","금천구":"101080",
    "노원구":"101090","도봉구":"101100","동대문구":"101110",
    "동작구":"101120","마포구":"101130","서대문구":"101140",
    "서초구":"101150","성동구":"101160","성북구":"101170",
    "송파구":"101180","양천구":"101190","영등포구":"101200",
    "용산구":"101210","은평구":"101220","종로구":"101230",
    "중구(서울)":"101240","중랑구":"101250",
    # 경기
    "경기 전체":"102000","수원시":"102010","성남시":"102020",
    "의정부시":"102030","안양시":"102040","부천시":"102050",
    "광명시":"102060","평택시":"102080","안산시":"102100",
    "고양시":"102110","남양주시":"102130","화성시":"102150",
    "용인시":"102160","파주시":"102170","구리시":"102185",
    "하남시":"102195",
    # 인천
    "인천 전체":"103000","계양구":"103010","남동구":"103020",
    "동구(인천)":"103030","미추홀구":"103040","부평구":"103050",
    "서구(인천)":"103060","연수구":"103070","중구(인천)":"103080",
    # 부산
    "부산 전체":"104000","강서구(부산)":"104010","금정구":"104020",
    "기장군":"104030","남구(부산)":"104040","동구(부산)":"104050",
    "동래구":"104060","부산진구":"104070","북구(부산)":"104080",
    "사상구":"104090","사하구":"104100","서구(부산)":"104110",
    "수영구":"104120","연제구":"104130","영도구":"104140",
    "중구(부산)":"104150","해운대구":"104160",
    # 대구
    "대구 전체":"105000","달서구":"105010","달성군":"105020",
    "동구(대구)":"105030","북구(대구)":"105040","서구(대구)":"105050",
    "수성구":"105060","중구(대구)":"105070",
    # 광주
    "광주 전체":"106000","광산구":"106010","남구(광주)":"106020",
    "동구(광주)":"106030","북구(광주)":"106040","서구(광주)":"106050",
    # 대전
    "대전 전체":"107000","대덕구":"107010","동구(대전)":"107020",
    "서구(대전)":"107030","유성구":"107040","중구(대전)":"107050",
    # 울산
    "울산 전체":"108000","남구(울산)":"108010","동구(울산)":"108020",
    "북구(울산)":"108030","울주군":"108040","중구(울산)":"108050",
    # 세종
    "세종특별자치시":"109000",
    # 강원
    "강원 전체":"110000","강릉시":"110010","원주시":"110090",
    "춘천시":"110130","속초시":"110050","동해시":"110030",
    "삼척시":"110040","태백시":"110140","홍천군":"110160",
    "횡성군":"110180","영월군":"110080","평창군":"110150",
    "정선군":"110110","철원군":"110120","화천군":"110170",
    "양구군":"110060","인제군":"110100","고성군(강원)":"110020",
    "양양군":"110070",
    # 충북
    "충북 전체":"111000","청주시":"111100","충주시":"111110",
    "제천시":"111070","음성군":"111060","진천군":"111090",
    "증평군":"111080","괴산군":"111010","보은군":"111030",
    "옥천군":"111050","영동군":"111040","단양군":"111020",
    # 충남
    "충남 전체":"112000","천안시":"112120","아산시":"112100",
    "서산시":"112080","당진시":"112050","공주시":"112020",
    "보령시":"112060","논산시":"112040","계룡시":"112010",
    "홍성군":"112150","예산군":"112110","부여군":"112070",
    "서천군":"112090","청양군":"112130","태안군":"112140",
    "금산군":"112030",
    # 전북
    "전북 전체":"113000","전주시":"113120","익산시":"113090",
    "군산시":"113020","정읍시":"113130","남원시":"113040",
    "김제시":"113030","완주군":"113080","고창군":"113010",
    "부안군":"113060","임실군":"113100","순창군":"113070",
    "무주군":"113050","장수군":"113110","진안군":"113140",
    # 전남
    "전남 전체":"114000","목포시":"114080","여수시":"114130",
    "순천시":"114110","나주시":"114060","광양시":"114040",
    "담양군":"114070","곡성군":"114030","구례군":"114050",
    "고흥군":"114020","보성군":"114100","화순군":"114220",
    "장흥군":"114180","강진군":"114010","해남군":"114210",
    "영암군":"114150","무안군":"114090","함평군":"114200",
    "영광군":"114140","장성군":"114170","완도군":"114160",
    "진도군":"114190","신안군":"114120",
    # 경북
    "경북 전체":"115000","포항시":"115230","경주시":"115020",
    "김천시":"115060","안동시":"115110","구미시":"115040",
    "영주시":"115140","영천시":"115150","상주시":"115090",
    "문경시":"115070","경산시":"115010","군위군":"115050",
    "의성군":"115190","청송군":"115210","영양군":"115130",
    "영덕군":"115120","청도군":"115200","고령군":"115030",
    "성주군":"115100","칠곡군":"115220","예천군":"115160",
    "봉화군":"115080","울진군":"115180","울릉군":"115170",
    # 경남
    "경남 전체":"116000","창원시":"116130","진주시":"116110",
    "통영시":"116140","사천시":"116070","김해시":"116040",
    "밀양시":"116060","거제시":"116010","양산시":"116090",
    "의령군":"116100","함안군":"116160","창녕군":"116120",
    "고성군(경남)":"116030","남해군":"116050","하동군":"116150",
    "산청군":"116080","함양군":"116170","거창군":"116020",
    "합천군":"116180",
    # 제주
    "제주 전체":"117000","제주시":"117020","서귀포시":"117010",
}

# 지역 그룹 (콤보박스용)
LOC_GROUPS = [
    ("전국", ["전국"]),
    ("서울", ["서울 전체","강남구","강동구","강북구","강서구","관악구","광진구",
              "구로구","금천구","노원구","도봉구","동대문구","동작구","마포구",
              "서대문구","서초구","성동구","성북구","송파구","양천구","영등포구",
              "용산구","은평구","종로구","중구(서울)","중랑구"]),
    ("경기", ["경기 전체","수원시","성남시","의정부시","안양시","부천시",
              "광명시","평택시","안산시","고양시","남양주시","화성시",
              "용인시","파주시","구리시","하남시"]),
    ("인천", ["인천 전체","계양구","남동구","동구(인천)","미추홀구",
              "부평구","서구(인천)","연수구","중구(인천)"]),
    ("부산", ["부산 전체","강서구(부산)","금정구","기장군","남구(부산)",
              "동구(부산)","동래구","부산진구","북구(부산)","사상구",
              "사하구","서구(부산)","수영구","연제구","영도구",
              "중구(부산)","해운대구"]),
    ("대구", ["대구 전체","달서구","달성군","동구(대구)","북구(대구)",
              "서구(대구)","수성구","중구(대구)"]),
    ("광주", ["광주 전체","광산구","남구(광주)","동구(광주)","북구(광주)","서구(광주)"]),
    ("대전", ["대전 전체","대덕구","동구(대전)","서구(대전)","유성구","중구(대전)"]),
    ("울산", ["울산 전체","남구(울산)","동구(울산)","북구(울산)","울주군","중구(울산)"]),
    ("세종", ["세종특별자치시"]),
    ("강원", ["강원 전체","강릉시","원주시","춘천시","속초시","동해시",
              "삼척시","태백시","홍천군","횡성군","영월군","평창군",
              "정선군","철원군","화천군","양구군","인제군","고성군(강원)","양양군"]),
    ("충북", ["충북 전체","청주시","충주시","제천시","음성군","진천군",
              "증평군","괴산군","보은군","옥천군","영동군","단양군"]),
    ("충남", ["충남 전체","천안시","아산시","서산시","당진시","공주시",
              "보령시","논산시","계룡시","홍성군","예산군","부여군",
              "서천군","청양군","태안군","금산군"]),
    ("전북", ["전북 전체","전주시","익산시","군산시","정읍시","남원시",
              "김제시","완주군","고창군","부안군","임실군","순창군",
              "무주군","장수군","진안군"]),
    ("전남", ["전남 전체","목포시","여수시","순천시","나주시","광양시",
              "담양군","곡성군","구례군","고흥군","보성군","화순군",
              "장흥군","강진군","해남군","영암군","무안군","함평군",
              "영광군","장성군","완도군","진도군","신안군"]),
    ("경북", ["경북 전체","포항시","경주시","김천시","안동시","구미시",
              "영주시","영천시","상주시","문경시","경산시","군위군",
              "의성군","청송군","영양군","영덕군","청도군","고령군",
              "성주군","칠곡군","예천군","봉화군","울진군","울릉군"]),
    ("경남", ["경남 전체","창원시","진주시","통영시","사천시","김해시",
              "밀양시","거제시","양산시","의령군","함안군","창녕군",
              "고성군(경남)","남해군","하동군","산청군","함양군","거창군","합천군"]),
    ("제주", ["제주 전체","제주시","서귀포시"]),
]

JT_MAP  = {"전체":"","정규직":"1","계약직":"2","인턴":"3","알바":"4"}
EXP_MAP = {"전체":"","신입":"1","경력":"2","신입·경력":"3"}
SAL_MAP = {"전체":"","회사내규":"1","2400만↑":"2","3000만↑":"3",
           "3600만↑":"4","4200만↑":"5"}
SRT_MAP = {"최신순":"pd","마감순":"pt","관련도":"sc","급여순":"sm"}

# ══════════════════════════════════════
# 유틸
# ══════════════════════════════════════
def get_internet_date():
    servers = [
        "https://worldtimeapi.org/api/timezone/Asia/Seoul",
        "https://timeapi.io/api/Time/current/zone?timeZone=Asia/Seoul",
    ]
    for url in servers:
        try:
            req = urllib.request.Request(url,headers={"User-Agent":"Mozilla/5.0"})
            with urllib.request.urlopen(req,timeout=6) as r:
                d=json.loads(r.read())
                dt=d.get("datetime") or d.get("dateTime","")
                if dt: return datetime.date.fromisoformat(dt[:10])
        except: continue
    return None

def fmt_date(ts):
    if not ts: return "상시채용"
    return datetime.datetime.fromtimestamp(int(ts)).strftime("%Y.%m.%d")

def exp_str(elv):
    if not elv: return "경력무관"
    nm=elv.get("name","")
    if nm: return nm
    mn,mx=elv.get("min",0) or 0,elv.get("max",0) or 0
    if not mn and not mx: return "경력무관"
    if not mn: return "신입"
    return f"경력 {mn}년↑"

def sal_str(code):
    return{"1":"회사내규","2":"2400만↑","3":"3000만↑",
           "4":"3600만↑","5":"4200만↑","6":"5000만↑"}.get(str(code or ""),"면접후결정")

def jt_str(code):
    return{"1":"정규직","2":"계약직","3":"인턴","4":"알바"}.get(str(code or ""),"-")

# ══════════════════════════════════════
# 비밀번호 다이얼로그
# ══════════════════════════════════════
def show_password_dialog():
    result={"ok":False}
    dlg=tk.Tk()
    dlg.title("채용정보 검색기 — 비밀번호")
    dlg.resizable(False,False)
    dlg.configure(bg="#1A3A6B")
    w,h=380,210
    sw,sh=dlg.winfo_screenwidth(),dlg.winfo_screenheight()
    dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    tk.Label(dlg,text="💼 채용정보 검색기",font=("맑은 고딕",14,"bold"),
             bg="#1A3A6B",fg="white").pack(pady=(22,4))
    tk.Label(dlg,text="비밀번호를 입력하세요",font=("맑은 고딕",10),
             bg="#1A3A6B",fg="#93C5FD").pack()

    frm=tk.Frame(dlg,bg="#1A3A6B"); frm.pack(pady=14)
    entry=tk.Entry(frm,show="●",font=("맑은 고딕",13),width=20,
                   relief="flat",bd=6,bg="#EFF6FF")
    entry.pack(); entry.focus_set()

    def on_ok(e=None):
        if entry.get()==PASSWORD:
            result["ok"]=True; dlg.destroy()
        else:
            messagebox.showerror("오류","비밀번호가 틀렸습니다.",parent=dlg)
            entry.delete(0,tk.END)

    entry.bind("<Return>",on_ok)
    dlg.protocol("WM_DELETE_WINDOW",dlg.destroy)

    btn=tk.Frame(dlg,bg="#1A3A6B"); btn.pack()
    tk.Button(btn,text="  확인  ",font=("맑은 고딕",10,"bold"),
              bg="white",fg="#1A3A6B",relief="flat",cursor="hand2",
              command=on_ok).pack(side="left",padx=8)
    tk.Button(btn,text="  취소  ",font=("맑은 고딕",10),
              bg="#6B7280",fg="white",relief="flat",cursor="hand2",
              command=dlg.destroy).pack(side="left",padx=8)

    tk.Label(dlg,text="50+재단 동부캠퍼스 | 정병남 취업코디",
             font=("맑은 고딕",8),bg="#1A3A6B",fg="#6B9FD4").pack(pady=(12,0))
    dlg.mainloop()
    return result["ok"]

# ══════════════════════════════════════
# 메인 앱
# ══════════════════════════════════════
class App:
    def __init__(self,root):
        self.root=root
        self.root.title(APP_TITLE)
        self.root.geometry("1340x860")
        self.root.minsize(1100,700)
        self.root.configure(bg="#F0F2F5")

        # 검색 상태
        self.jobs=[]
        self.cur_page=1
        self.total_pages=1
        self.total_count=0
        self.total_req=0
        self.batch_start=1
        self.search_params={}

        self._build_ui()
        self.root.after(200,self._show_notice)

    def _show_notice(self):
        messagebox.showinfo("이용 안내",NOTICE_MSG,parent=self.root)

    # ── UI 구성 ──
    def _build_ui(self):
        self._build_header()
        self._build_search()
        self._build_result_bar()
        self._build_table()
        self._build_pagination()
        self._build_statusbar()
        self._apply_style()

    def _build_header(self):
        hdr=tk.Frame(self.root,bg="#1A56A0",height=58)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr,text="💼  채용정보 검색기 — 50+재단 동부캠퍼스",
                 font=("맑은 고딕",16,"bold"),
                 bg="#1A56A0",fg="white").pack(side="left",padx=20)
        tk.Label(hdr,text="작성자: 정병남 취업코디  |  Powered by 사람인 API",
                 font=("맑은 고딕",9),bg="#1A56A0",fg="#93C5FD").pack(side="right",padx=20)

    def _build_search(self):
        panel=tk.Frame(self.root,bg="white",pady=10)
        panel.pack(fill="x")

        # 키워드 행
        kw_row=tk.Frame(panel,bg="white")
        kw_row.pack(fill="x",padx=16,pady=(0,8))
        tk.Label(kw_row,text="검색 키워드",font=("맑은 고딕",10,"bold"),
                 bg="white",fg="#374151").pack(side="left")
        self.kw_var=tk.StringVar()
        kw=tk.Entry(kw_row,textvariable=self.kw_var,
                    font=("맑은 고딕",12),width=40,
                    relief="solid",bd=1,bg="#F9FAFB")
        kw.pack(side="left",padx=(8,12),ipady=6)
        kw.bind("<Return>",lambda e:self.do_search())
        self.srch_btn=tk.Button(kw_row,text="🔍  검색",
                                 font=("맑은 고딕",11,"bold"),
                                 bg="#1A56A0",fg="white",
                                 relief="flat",padx=18,pady=7,
                                 cursor="hand2",command=self.do_search)
        self.srch_btn.pack(side="left")

        tk.Frame(panel,bg="#E5E7EB",height=1).pack(fill="x",padx=16,pady=4)

        # 필터 행
        f_row=tk.Frame(panel,bg="white")
        f_row.pack(fill="x",padx=16)

        # 지역 — 그룹별 Combobox
        loc_box=tk.Frame(f_row,bg="white"); loc_box.pack(side="left",padx=8)
        tk.Label(loc_box,text="지역",font=("맑은 고딕",9,"bold"),
                 bg="white",fg="#6B7280").pack(anchor="w")
        self.loc_var=tk.StringVar(value="전국")
        all_locs=[]
        for grp,items in LOC_GROUPS:
            all_locs.extend(items)
        loc_cb=ttk.Combobox(loc_box,textvariable=self.loc_var,
                             values=all_locs,width=14,state="readonly",
                             font=("맑은 고딕",9))
        loc_cb.pack()

        # 나머지 필터
        others=[
            ("고용형태","jt",list(JT_MAP.keys()),9),
            ("경력","exp",list(EXP_MAP.keys()),9),
            ("연봉","sal",list(SAL_MAP.keys()),9),
            ("정렬","srt",list(SRT_MAP.keys()),9),
            ("검색건수","cnt",
             ["10건","20건","30건","50건","100건","200건",
              "300건","500건","1000건","2000건","3000건"],10),
        ]
        self.fvars={}
        for label,key,opts,w in others:
            box=tk.Frame(f_row,bg="white"); box.pack(side="left",padx=8)
            tk.Label(box,text=label,font=("맑은 고딕",9,"bold"),
                     bg="white",fg="#6B7280").pack(anchor="w")
            var=tk.StringVar(value=opts[0])
            self.fvars[key]=var
            ttk.Combobox(box,textvariable=var,values=opts,
                         width=w,state="readonly",
                         font=("맑은 고딕",9)).pack()

    def _build_result_bar(self):
        bar=tk.Frame(self.root,bg="white",pady=5)
        bar.pack(fill="x")
        tk.Frame(bar,bg="#E5E7EB",height=1).pack(fill="x")
        inner=tk.Frame(bar,bg="white"); inner.pack(fill="x",padx=16,pady=4)
        self.stat_var=tk.StringVar(value="검색 결과가 여기에 표시됩니다.")
        tk.Label(inner,textvariable=self.stat_var,
                 font=("맑은 고딕",10),bg="white",fg="#374151").pack(side="left")
        tk.Button(inner,text="📊  엑셀 저장",
                  font=("맑은 고딕",9,"bold"),
                  bg="#2D7A3A",fg="white",relief="flat",
                  padx=12,pady=4,cursor="hand2",
                  command=self.do_excel).pack(side="right",padx=4)

    def _build_table(self):
        wrap=tk.Frame(self.root,bg="#F0F2F5")
        wrap.pack(fill="both",expand=True,padx=10,pady=6)
        cols=("연번","기업명","채용직무 (공고제목)","고용형태",
              "경력","근무지역","연봉/급여","마감일")
        self.tree=ttk.Treeview(wrap,columns=cols,
                                show="headings",selectmode="browse")
        widths=[44,155,330,78,90,140,90,96]
        for col,w in zip(cols,widths):
            anchor="center" if col in ("연번","고용형태","경력","연봉/급여","마감일") else "w"
            self.tree.heading(col,text=col)
            self.tree.column(col,width=w,minwidth=40,anchor=anchor)
        vsb=ttk.Scrollbar(wrap,orient="vertical",command=self.tree.yview)
        hsb=ttk.Scrollbar(wrap,orient="horizontal",command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set,xscrollcommand=hsb.set)
        self.tree.grid(row=0,column=0,sticky="nsew")
        vsb.grid(row=0,column=1,sticky="ns")
        hsb.grid(row=1,column=0,sticky="ew")
        wrap.grid_rowconfigure(0,weight=1)
        wrap.grid_columnconfigure(0,weight=1)
        self.tree.bind("<Double-1>",self._on_detail)
        self.tree.bind("<Return>",self._on_detail)
        self.tree.tag_configure("odd",background="#F9FAFB")
        self.tree.tag_configure("even",background="#FFFFFF")
        self.tree.tag_configure("warn",background="#FFF1F2",foreground="#B91C1C")

    def _build_pagination(self):
        self.pg_frame=tk.Frame(self.root,bg="#F0F2F5")
        self.pg_frame.pack(pady=4)

    def _build_statusbar(self):
        self.status_var=tk.StringVar(value="키워드와 조건을 입력하고 검색 버튼을 누르세요.")
        tk.Label(self.root,textvariable=self.status_var,
                 font=("맑은 고딕",9),bg="#DDE1E9",
                 fg="#6B7280",anchor="w",padx=10).pack(fill="x",side="bottom")

    def _apply_style(self):
        s=ttk.Style(); s.theme_use("clam")
        s.configure("Treeview",font=("맑은 고딕",9),rowheight=26,
                    fieldbackground="#FFFFFF")
        s.configure("Treeview.Heading",font=("맑은 고딕",9,"bold"),
                    background="#F1F5F9",foreground="#374151")
        s.map("Treeview",background=[("selected","#DBEAFE")],
              foreground=[("selected","#1E40AF")])

    # ══ 검색 ══
    def do_search(self):
        kw=self.kw_var.get().strip()
        cnt_raw=self.fvars["cnt"].get().replace("건","")
        self.total_req=min(int(cnt_raw),3000)

        # 파라미터 저장
        self.search_params={
            "kw":kw,
            "loc":LOC_MAP.get(self.loc_var.get(),""),
            "jt":JT_MAP.get(self.fvars["jt"].get(),""),
            "exp":EXP_MAP.get(self.fvars["exp"].get(),""),
            "sal":SAL_MAP.get(self.fvars["sal"].get(),""),
            "srt":SRT_MAP.get(self.fvars["srt"].get(),"pd"),
        }

        self.jobs=[]
        self.batch_start=1
        self.srch_btn.config(state="disabled",text="⏳  검색 중...")
        self.status_var.set("검색 중입니다...")
        self.tree.delete(*self.tree.get_children())
        self._clear_pg()

        cnt=min(self.total_req,BATCH_SIZE)
        threading.Thread(target=self._fetch_batch,
                         args=(1,cnt,True),daemon=True).start()

    def _fetch_batch(self,start,cnt,is_first):
        p=self.search_params
        params={"access-key":API_KEY,"count":cnt,"start":start,"sort":p["srt"]}
        if p["kw"]:  params["keywords"]=p["kw"]
        if p["loc"]: params["loc_cd"]=p["loc"]
        if p["jt"]:  params["job_type"]=p["jt"]
        if p["exp"]: params["exp_cd"]=p["exp"]
        if p["sal"]: params["sal_cd"]=p["sal"]

        try:
            url=PROXY_URL+"?"+urllib.parse.urlencode(params)
            req=urllib.request.Request(url,headers={
                "Accept":"application/json","User-Agent":"Mozilla/5.0"})
            with urllib.request.urlopen(req,timeout=15) as r:
                data=json.loads(r.read())

            jd=data.get("jobs",{})
            if is_first:
                self.total_count=int(jd.get("total",0))
            job_list=jd.get("job",[])
            if isinstance(job_list,dict): job_list=[job_list]

            self.jobs.extend(job_list)
            self.batch_start=start+len(job_list)
            self.root.after(0,lambda:self._on_batch_done(is_first))
        except Exception as e:
            self.root.after(0,lambda:self._on_error(str(e)))

    def _on_batch_done(self,is_first):
        self.srch_btn.config(state="normal",text="🔍  검색")
        if not self.jobs:
            self.status_var.set("검색 결과가 없습니다.")
            messagebox.showinfo("검색 결과","검색 결과가 없습니다.\n조건을 변경해보세요.",
                                parent=self.root)
            return
        self.total_pages=max(1,(len(self.jobs)+PAGE_SIZE-1)//PAGE_SIZE)
        self.cur_page=1
        self._show_page(1)
        loaded=len(self.jobs)
        remain=min(self.total_req,self.total_count)-loaded
        self.status_var.set(
            f"총 {self.total_count:,}건 검색 완료 · "
            f"{loaded}건 로드 완료"
            +(f" · 남은 {remain}건 더보기 가능" if remain>0 else ""))

    def _on_error(self,msg):
        self.srch_btn.config(state="normal",text="🔍  검색")
        self.status_var.set(f"오류: {msg}")
        messagebox.showerror("검색 오류",
            f"검색 중 오류가 발생했습니다.\n\n{msg}\n\n인터넷 연결을 확인해주세요.",
            parent=self.root)

    # ★ 다음 배치 로드
    def load_next(self):
        remain=min(self.total_req,self.total_count)-len(self.jobs)
        if remain<=0:
            messagebox.showinfo("알림","더 이상 불러올 데이터가 없습니다.",
                                parent=self.root); return
        cnt=min(BATCH_SIZE,remain)
        self.srch_btn.config(state="disabled",text="⏳  검색 중...")
        self.status_var.set(f"{self.batch_start}번부터 {cnt}건 추가 로드 중...")
        threading.Thread(target=self._fetch_batch,
                         args=(self.batch_start,cnt,False),daemon=True).start()

    # ══ 페이지 표시 ══
    def _show_page(self,p):
        self.cur_page=p
        self.tree.delete(*self.tree.get_children())
        start=(p-1)*PAGE_SIZE
        page=self.jobs[start:start+PAGE_SIZE]
        now=datetime.datetime.now()
        for i,v in enumerate(page):
            corp=((v.get("company") or {}).get("detail") or {}).get("name","-")
            title=v.get("position",{}).get("title","-")
            loc=((v.get("position") or {}).get("location") or {}).get("name","-")
            loc=loc.replace("&gt;",">")
            jtc=((v.get("position") or {}).get("job-type") or {}).get("code","")
            elv=(v.get("position") or {}).get("experience-level") or {}
            sc=(v.get("salary") or {}).get("code","")
            ts=v.get("expiration-timestamp")
            dead=fmt_date(ts)
            diff=(datetime.datetime.fromtimestamp(int(ts))-now).days if ts else 999
            tag="warn" if ts and diff<=7 else ("odd" if i%2==1 else "even")
            self.tree.insert("","end",iid=str(start+i),
                values=(start+i+1,corp,title,jt_str(jtc),
                        exp_str(elv),loc,sal_str(sc),dead),tags=(tag,))
        self.stat_var.set(
            f"총 {self.total_count:,}건 · {len(page)}건 표시 "
            f"(페이지 {self.cur_page}/{self.total_pages})")
        self._render_pg()

    # ══ 페이지네이션 ══
    def _clear_pg(self):
        for w in self.pg_frame.winfo_children(): w.destroy()

    def _render_pg(self):
        self._clear_pg()
        def mk(label,page,active=False,disabled=False):
            b=tk.Button(self.pg_frame,text=label,
                        font=("맑은 고딕",9,"bold" if active else "normal"),
                        width=3,pady=2,
                        bg="#1A56A0" if active else "#FFFFFF",
                        fg="white" if active else "#6B7280",
                        relief="solid",bd=1,
                        state="disabled" if disabled else "normal",
                        cursor="arrow" if disabled else "hand2",
                        command=lambda p=page:self._show_page(p))
            b.pack(side="left",padx=2)

        p=self.cur_page
        if self.total_pages>1:
            mk("«",1,disabled=(p==1))
            mk("‹",p-1,disabled=(p==1))
            s=max(1,p-2); e=min(self.total_pages,s+4)
            if e-s<4: s=max(1,e-4)
            for i in range(s,e+1): mk(str(i),i,active=(i==p))
            mk("›",p+1,disabled=(p==self.total_pages))
            mk("»",self.total_pages,disabled=(p==self.total_pages))

        # ★ 다음 배치 버튼
        remain=min(self.total_req,self.total_count)-len(self.jobs)
        if remain>0 and self.batch_start>1:
            sep=tk.Label(self.pg_frame,text=" | ",
                         font=("맑은 고딕",12),bg="#F0F2F5",fg="#9CA3AF")
            sep.pack(side="left")
            next_cnt=min(BATCH_SIZE,remain)
            nb=tk.Button(self.pg_frame,
                         text=f"다음 {next_cnt}건 ▶",
                         font=("맑은 고딕",9,"bold"),
                         pady=2,padx=10,
                         bg="#2D7A3A",fg="white",
                         relief="solid",bd=1,cursor="hand2",
                         command=self.load_next)
            nb.pack(side="left",padx=2)

    # ══ 상세보기 ══
    def _on_detail(self,event):
        sel=self.tree.selection()
        if not sel: return
        idx=int(sel[0])
        if idx>=len(self.jobs): return
        v=self.jobs[idx]
        corp=((v.get("company") or {}).get("detail") or {}).get("name","-")
        title=v.get("position",{}).get("title","-")
        loc=((v.get("position") or {}).get("location") or {}).get("name","-").replace("&gt;",">")
        jtn=((v.get("position") or {}).get("job-type") or {}).get("name","-")
        elv=(v.get("position") or {}).get("experience-level") or {}
        sc=(v.get("salary") or {}).get("code","")
        edu=((v.get("position") or {}).get("required-education-level") or {}).get("name","학력무관")
        dead=fmt_date(v.get("expiration-timestamp"))
        ind=((v.get("position") or {}).get("industry") or {}).get("name","-")
        url=v.get("url","")

        win=tk.Toplevel(self.root)
        win.title(f"상세보기 — {corp}")
        win.geometry("640x500")
        win.configure(bg="#F9FAFB")
        win.grab_set()

        hdr=tk.Frame(win,bg="#1A56A0",height=56)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr,text=corp,font=("맑은 고딕",13,"bold"),
                 bg="#1A56A0",fg="white").pack(side="left",padx=16)

        body=tk.Frame(win,bg="#F9FAFB")
        body.pack(fill="both",expand=True,padx=20,pady=14)

        tk.Label(body,text=title,font=("맑은 고딕",14,"bold"),
                 bg="#F9FAFB",fg="#1A1D23",
                 wraplength=590,justify="left").pack(anchor="w",pady=(0,12))

        rows=[("직종분류",ind),("고용형태",jtn),("경력조건",exp_str(elv)),
              ("학력",edu),("연봉/급여",sal_str(sc)),("근무지역",loc),("마감일",dead)]
        for label,val in rows:
            r=tk.Frame(body,bg="#FFFFFF",
                       highlightbackground="#E5E7EB",highlightthickness=1)
            r.pack(fill="x",pady=2)
            tk.Label(r,text=label,font=("맑은 고딕",9,"bold"),
                     width=10,bg="#F3F4F6",fg="#6B7280",
                     pady=6,anchor="center").pack(side="left")
            tk.Label(r,text=val,font=("맑은 고딕",10),
                     bg="#FFFFFF",fg="#1A1D23",
                     anchor="w",padx=10).pack(side="left")

        tk.Label(body,text="채용공고 URL",font=("맑은 고딕",9,"bold"),
                 bg="#F9FAFB",fg="#6B7280").pack(anchor="w",pady=(10,2))
        ue=tk.Entry(body,font=("맑은 고딕",9),relief="solid",bd=1,bg="#F3F4F6")
        ue.insert(0,url); ue.config(state="readonly"); ue.pack(fill="x")

        br=tk.Frame(win,bg="#F9FAFB"); br.pack(pady=10)
        tk.Button(br,text="🔗  브라우저에서 공고 열기",
                  font=("맑은 고딕",11,"bold"),bg="#1A56A0",fg="white",
                  relief="flat",padx=20,pady=8,cursor="hand2",
                  command=lambda:webbrowser.open(url)).pack(side="left",padx=8)
        tk.Button(br,text="닫기",font=("맑은 고딕",10),
                  bg="#6B7280",fg="white",relief="flat",
                  padx=16,pady=8,cursor="hand2",
                  command=win.destroy).pack(side="left")

    # ══ 엑셀 저장 ══
    def do_excel(self):
        if not self.jobs:
            messagebox.showwarning("알림","저장할 데이터가 없습니다.",parent=self.root); return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
            from openpyxl.utils import get_column_letter

            today=datetime.date.today().strftime("%Y%m%d")
            kw=self.kw_var.get().strip() or "전체"
            fname=filedialog.asksaveasfilename(
                parent=self.root,defaultextension=".xlsx",
                filetypes=[("Excel 파일","*.xlsx")],
                initialfile=f"사람인_채용정보_{kw}_{today}.xlsx")
            if not fname: return

            fl=lambda rgb:PatternFill("solid",fgColor=rgb)
            fn=lambda bold=False,sz=9,color="000000",ul=None:Font(
                bold=bold,size=sz,color=color,name="맑은 고딕",underline=ul)
            al=lambda h="left",v="center",wrap=False:Alignment(
                horizontal=h,vertical=v,wrap_text=wrap)
            bd=lambda:Border(*[Side(style="thin",color="CCCCCC")]*0,
                             **{k:Side(style="thin",color="CCCCCC")
                                for k in["left","right","top","bottom"]})

            wb=Workbook(); ws=wb.active; ws.title="채용정보"
            COLS=12; today_s=datetime.date.today().strftime("%Y.%m.%d")

            ws.row_dimensions[1].height=28
            ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
            c=ws["A1"]
            c.value=f"사람인 채용정보 검색결과 — {kw} | 검색일: {today_s}"
            c.fill=fl("1A3A2A"); c.font=fn(bold=True,sz=13,color="FFFFFF")
            c.alignment=al("left","center")

            ws.row_dimensions[2].height=16
            ws.merge_cells(f"A2:{get_column_letter(COLS)}2")
            c=ws["A2"]
            c.value=(f"■ 총 {self.total_count}건 중 "
                     f"{len(self.jobs)}건 | 출처: 사람인(saramin.co.kr)")
            c.fill=fl("DFF0E3"); c.font=fn(sz=8,color="1A3A2A")
            c.alignment=al("left","center")

            ws.row_dimensions[3].height=30
            headers=["연번","기업명","채용 직무(공고제목)","직종분류","고용형태",
                     "경력조건","근무지역","연봉/급여","마감일","채용공고 URL","등록일","코디 메모"]
            for ci,h in enumerate(headers,1):
                c=ws.cell(row=3,column=ci,value=h)
                c.fill=fl("1E6B3A"); c.font=fn(bold=True,sz=9,color="FFFFFF")
                c.alignment=al("center","center",wrap=True); c.border=bd()

            for i,w in enumerate([5,18,30,14,9,11,16,11,11,28,9,24],1):
                ws.column_dimensions[get_column_letter(i)].width=w

            for ri,v in enumerate(self.jobs,1):
                rn=ri+3; bg="F5F8F5" if ri%2==1 else "FFFFFF"
                ws.row_dimensions[rn].height=26
                corp=((v.get("company") or {}).get("detail") or {}).get("name","-")
                title=v.get("position",{}).get("title","-")
                loc=((v.get("position") or {}).get("location") or {}).get("name","-").replace("&gt;",">")
                jtc=((v.get("position") or {}).get("job-type") or {}).get("code","")
                elv=(v.get("position") or {}).get("experience-level") or {}
                ind=((v.get("position") or {}).get("industry") or {}).get("name","-")
                sc=(v.get("salary") or {}).get("code","")
                dead=fmt_date(v.get("expiration-timestamp"))
                url=v.get("url","")
                ts2=v.get("posting-timestamp")
                reg=(datetime.datetime.fromtimestamp(int(ts2)).strftime("%Y-%m")
                     if ts2 else today_s[:7])

                vals=[ri,corp,title,ind,jt_str(jtc),exp_str(elv),
                      loc,sal_str(sc),dead,None,reg,""]
                hc={1,5,6,8,9,11}
                for ci,val in enumerate(vals,1):
                    if ci==10: continue
                    c=ws.cell(row=rn,column=ci,value=val)
                    c.fill=fl(bg); c.font=fn(bold=(ci==2),sz=9)
                    c.alignment=al("center" if ci in hc else "left",
                                   "center",wrap=(ci in{3,7,12}))
                    c.border=bd()

                uc=ws.cell(row=rn,column=10)
                if url:
                    uc.value="공고 바로가기"; uc.hyperlink=url
                    uc.fill=fl("E8F0FB")
                    uc.font=Font(bold=True,size=9,color="1A56A0",
                                 underline="single",name="맑은 고딕")
                    uc.alignment=al("center","center"); uc.border=bd()
                else:
                    uc.value="-"; uc.fill=fl(bg); uc.font=fn(sz=9)
                    uc.alignment=al("center","center"); uc.border=bd()

            ws.freeze_panes="A4"
            wb.save(fname)
            messagebox.showinfo("저장 완료",f"엑셀 파일이 저장되었습니다!\n\n{fname}",
                                parent=self.root)
        except ImportError:
            messagebox.showerror("오류","openpyxl이 필요합니다.\npip install openpyxl",
                                 parent=self.root)
        except Exception as e:
            messagebox.showerror("저장 오류",str(e),parent=self.root)


# ══════════════════════════════════════
# 진입점
# ══════════════════════════════════════
def main():
    # ① 유효기간 확인
    tmp=tk.Tk(); tmp.withdraw()
    today=get_internet_date()
    if today is None:
        messagebox.showerror("인터넷 연결 필요",
            "유효기간 확인을 위해 인터넷 연결이 필요합니다.\n"
            "인터넷 연결 후 다시 실행해주세요.",parent=tmp)
        tmp.destroy(); sys.exit(0)
    if today>EXPIRE_DATE:
        messagebox.showerror("사용 기간 만료",
            f"이 프로그램의 사용 기간이 만료되었습니다.\n"
            f"만료일: {EXPIRE_DATE.strftime('%Y년 %m월 %d일')}\n\n"
            "담당자에게 문의해주세요.",parent=tmp)
        tmp.destroy(); sys.exit(0)
    tmp.destroy()

    # ② 비밀번호 확인
    if not show_password_dialog(): sys.exit(0)

    # ③ 메인 앱
    root=tk.Tk()
    App(root)
    root.mainloop()

if __name__=="__main__":
    main()
